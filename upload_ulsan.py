# 실행 전 패키지 설치:
# pip install openpyxl supabase python-dotenv --break-system-packages
#
# 실행 방법:
# python upload_ulsan.py

import os
import sys
import time
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client, Client


SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = SCRIPT_DIR / "울산주소.xlsx"
ENV_PATH = SCRIPT_DIR / ".env.local"

BATCH_SIZE = 500
BATCH_SLEEP_SECONDS = 0.5

DEFAULT_NAME = None
DEFAULT_PASSWORD = None
DEFAULT_ACCESS_TYPE = "password"
DEFAULT_REGION = "울산"
DEFAULT_BRANCH_ID = "sinjeong"


def load_env() -> tuple[str, str]:
    if not ENV_PATH.exists():
        sys.exit(f".env.local 파일을 찾을 수 없습니다: {ENV_PATH}")
    load_dotenv(ENV_PATH)
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit(
            "환경변수 누락: NEXT_PUBLIC_SUPABASE_URL 또는 SUPABASE_SERVICE_ROLE_KEY"
        )
    return url, key


def read_excel(path: Path) -> list[dict]:
    if not path.exists():
        sys.exit(f"엑셀 파일을 찾을 수 없습니다: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        sys.exit("엑셀 워크시트를 읽을 수 없습니다.")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        sys.exit("엑셀이 비어 있습니다.")

    header_map = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
    for required in ("address", "lat", "lng"):
        if required not in header_map:
            sys.exit(f"엑셀에 '{required}' 컬럼이 없습니다. 헤더: {header}")

    addr_idx = header_map["address"]
    lat_idx = header_map["lat"]
    lng_idx = header_map["lng"]

    rows: list[dict] = []
    for r in rows_iter:
        if not r:
            continue
        address = r[addr_idx] if addr_idx < len(r) else None
        lat_raw = r[lat_idx] if lat_idx < len(r) else None
        lng_raw = r[lng_idx] if lng_idx < len(r) else None
        if address is None:
            continue
        address_s = str(address).strip()
        if not address_s:
            continue
        try:
            lat = float(lat_raw) if lat_raw is not None and str(lat_raw).strip() != "" else None
            lng = float(lng_raw) if lng_raw is not None and str(lng_raw).strip() != "" else None
        except (TypeError, ValueError):
            continue
        if lat is None or lng is None:
            continue
        rows.append({"address": address_s, "lat": lat, "lng": lng})

    wb.close()
    return rows


def fetch_existing(
    supabase: Client, addresses: list[str]
) -> dict[str, tuple[int, float | None, float | None]]:
    """name IS NULL인 기존 행을 가져온다. address → (id, lat, lng)."""
    existing: dict[str, tuple[int, float | None, float | None]] = {}
    chunk = 50
    for i in range(0, len(addresses), chunk):
        slice_ = addresses[i : i + chunk]
        resp = (
            supabase.table("buildings")
            .select("id, address, lat, lng")
            .in_("address", slice_)
            .is_("name", "null")
            .execute()
        )
        for row in resp.data or []:
            addr = row.get("address")
            rid = row.get("id")
            if addr and rid is not None and addr not in existing:
                existing[addr] = (rid, row.get("lat"), row.get("lng"))
    return existing


def _coords_equal(a: float | None, b: float | None) -> bool:
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) < 1e-9


def process_batch(
    supabase: Client, batch: list[dict]
) -> tuple[int, int]:
    addresses = [b["address"] for b in batch]
    existing = fetch_existing(supabase, addresses)

    to_insert: list[dict] = []
    inserted_count = 0
    updated_count = 0

    for row in batch:
        addr = row["address"]
        if addr in existing:
            rid, cur_lat, cur_lng = existing[addr]
            if _coords_equal(cur_lat, row["lat"]) and _coords_equal(cur_lng, row["lng"]):
                continue
            supabase.table("buildings").update(
                {"lat": row["lat"], "lng": row["lng"]}
            ).eq("id", rid).execute()
            updated_count += 1
        else:
            to_insert.append(
                {
                    "name": DEFAULT_NAME,
                    "address": addr,
                    "password": DEFAULT_PASSWORD,
                    "lat": row["lat"],
                    "lng": row["lng"],
                    "access_type": DEFAULT_ACCESS_TYPE,
                    "region": DEFAULT_REGION,
                    "branch_id": DEFAULT_BRANCH_ID,
                }
            )

    if to_insert:
        supabase.table("buildings").insert(to_insert).execute()
        inserted_count = len(to_insert)

    return inserted_count, updated_count


def format_elapsed(seconds: float) -> str:
    minutes = int(seconds // 60)
    secs = int(seconds % 60)
    return f"{minutes}분 {secs}초"


def main() -> None:
    url, key = load_env()
    supabase: Client = create_client(url, key)

    print(f"엑셀 로드: {EXCEL_PATH}")
    rows = read_excel(EXCEL_PATH)
    total = len(rows)
    if total == 0:
        sys.exit("엑셀에 처리할 데이터가 없습니다.")

    total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"총 {total:,}개 / 배치 {total_batches}개 (배치 크기 {BATCH_SIZE})")

    started = time.monotonic()
    total_inserted = 0
    total_updated = 0

    for i in range(total_batches):
        start = i * BATCH_SIZE
        end = min(start + BATCH_SIZE, total)
        batch = rows[start:end]
        inserted, updated = process_batch(supabase, batch)
        total_inserted += inserted
        total_updated += updated
        print(
            f"[{i + 1}/{total_batches}] 처리중 ({start}~{end})... "
            f"신규:{inserted} 업데이트:{updated}"
        )
        if i < total_batches - 1:
            time.sleep(BATCH_SLEEP_SECONDS)

    elapsed = time.monotonic() - started
    print("=" * 40)
    print(f"총 처리: {total:,}개")
    print(f"신규 추가: {total_inserted:,}개")
    print(f"업데이트: {total_updated:,}개")
    print(f"소요 시간: {format_elapsed(elapsed)}")
    print("=" * 40)


if __name__ == "__main__":
    main()
